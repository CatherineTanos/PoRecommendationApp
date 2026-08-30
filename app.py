"""
app.py
======
Streamlit dashboard for the PO recommendation engine - branded for
K. Beauty / The Beauty Shop.

Upload your 2 iPOS exports (penjualan 30 hari + stok) as-is - no manual
cleanup needed - and get an instant color-coded PO recommendation table,
per-brand sales & stock analytics, and a downloadable Excel file. No n8n,
no paid service, no external database.

RUN LOCALLY
-----------
    pip install -r requirements.txt
    streamlit run app.py

DEPLOY FREE (so your team can access it from a browser link)
--------------------------------------------------------------
    1. Push this folder to a GitHub repo (app.py, po_logic.py,
       requirements.txt, assets/logo.jpg, .streamlit/config.toml).
    2. Go to https://share.streamlit.io -> "New app" -> connect the repo.
    3. Streamlit Community Cloud builds and hosts it for free.
"""

import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from po_logic import (
    COLS,
    FLAG_COLORS,
    load_penjualan,
    load_stok,
    build_dataset,
    compute_recommendations,
)

LOGO_PATH = "assets/logo.jpg"
TEAL = "#186156"
LIME = "#ADF901"

st.set_page_config(
    page_title="K. Beauty - PO Recommendation Dashboard",
    page_icon=LOGO_PATH,
    layout="wide",
)

# ============================================================
# BRAND HEADER + THEME POLISH
# (base colors/backgrounds come from .streamlit/config.toml - this just
# adds a few accents that Streamlit's theme options don't cover)
# ============================================================
st.markdown(
    f"""
    <style>
    div[data-testid="stMetricValue"] {{ color: {TEAL}; }}
    h2, h3 {{ color: {TEAL}; }}
    .stTabs [aria-selected="true"] {{
        border-bottom: 3px solid {LIME} !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

logo_col, title_col = st.columns([1, 4])
with logo_col:
    st.image(LOGO_PATH, use_container_width=True)
with title_col:
    st.markdown(
        f"""
        <div style="padding-top: 22px;">
            <span style="font-size:26px; font-weight:700; color:{TEAL};">PO Recommendation Dashboard</span><br>
            <span style="font-size:15px; color:#5b8f83;">K. Beauty &middot; The Beauty Shop</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
st.markdown(
    f'<hr style="border:none; height:4px; background:linear-gradient(90deg,{TEAL},{LIME}); '
    'border-radius:4px; margin-top:6px; margin-bottom:18px;">',
    unsafe_allow_html=True,
)

# ============================================================
# SIDEBAR - inputs
# ============================================================

st.sidebar.image(LOGO_PATH, use_container_width=True)
st.sidebar.caption(
    "Upload export asli iPOS setiap minggu -> langsung dapat draft PO. "
    "Tidak perlu edit/rapikan file dulu."
)

penjualan_file = st.sidebar.file_uploader(
    "Data Penjualan 30 hari terakhir (Daftar Penjualan per Item)",
    type=["csv", "xlsx", "xls"],
)
stok_file = st.sidebar.file_uploader(
    "Data Stok (Daftar Item)", type=["csv", "xlsx", "xls"]
)

st.sidebar.divider()
st.sidebar.markdown(
    """
**Rumus yang dipakai:**
- `PO = Penjualan 30 hari − Stok`
- Stok == Penjualan 30 hari → 🟡 **CHECK SALES** (slow moving, PO tidak dibuat otomatis)
- Stok habis / hampir habis & masih laku → 🔴 **URGENT**
- Sisanya → 🟢 **NORMAL** (PO rutin jika PO_Qty > 0, atau stok masih cukup jika PO_Qty = 0)
    """
)

# ============================================================
# MAIN
# ============================================================

if not penjualan_file or not stok_file:
    st.info("Upload kedua file (penjualan & stok) di sidebar untuk mulai.")
    st.markdown(
        """
**Format file:** upload langsung hasil export iPOS 5.0 apa adanya
(*Daftar Penjualan per Item per Jenis* & *Daftar Item*/stok) - kolom
`Kode Item`, `Nama Item`, `Jenis`, `Jumlah`/`Stok` akan otomatis dibaca,
termasuk baris kosong dan baris "Total :" di antaranya.

Kalau suatu saat kamu punya file yang sudah rapi (1 baris = 1 header di
baris pertama), itu juga tetap bisa dibaca selama ada kolom `Product ID`/
`Kode Item`/`SKU` dan kolom qty/stok yang jelas.
        """
    )
    st.stop()

try:
    penjualan_df = load_penjualan(penjualan_file)
    stok_df = load_stok(stok_file)
except Exception as e:
    st.error(f"Gagal membaca file: {e}")
    st.stop()

missing = []
if COLS["product_id"] not in penjualan_df.columns:
    missing.append("Kolom Kode Item / Product ID tidak ditemukan di file penjualan")
if COLS["qty_terjual_30h"] not in penjualan_df.columns:
    missing.append("Kolom Jumlah / Qty Terjual tidak ditemukan di file penjualan")
if COLS["product_id"] not in stok_df.columns:
    missing.append("Kolom Kode Item / Product ID tidak ditemukan di file stok")
if COLS["stok"] not in stok_df.columns:
    missing.append("Kolom Stok tidak ditemukan di file stok")

if missing:
    st.error("Kolom tidak ditemukan:\n\n" + "\n".join(f"- {m}" for m in missing))
    st.stop()

df, dup_report = build_dataset(penjualan_df, stok_df)
df = compute_recommendations(df)

tab_current, tab_brand = st.tabs(["📊 Minggu Ini", "📦 Analitik per Brand"])

# ============================================================
# TAB 1: MINGGU INI
# ============================================================
with tab_current:

    total_dupes = dup_report["penjualan_duplicates_merged"] + dup_report["stok_duplicates_merged"]
    if total_dupes > 0:
        st.warning(
            f"🧹 Ditemukan & digabung {dup_report['penjualan_duplicates_merged']} baris duplikat di "
            f"data penjualan dan {dup_report['stok_duplicates_merged']} di data stok "
            f"(Product ID yang sama muncul lebih dari sekali - qty-nya sudah dijumlahkan)."
        )

    only_penjualan = int((df["Sumber Data"] == "Hanya ada di data penjualan (tidak ada di data stok)").sum())
    if only_penjualan > 0:
        st.info(
            f"ℹ️ {only_penjualan} SKU terjual dalam 30 hari terakhir tapi tidak ditemukan di data stok "
            f"(dianggap stok = 0 - cek apakah barang ini masih aktif/terdaftar)."
        )

    counts = df["Flag"].value_counts()
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total SKU", len(df))
    col2.metric("🔴 Urgent", int(counts.get("🔴 URGENT", 0)))
    col3.metric("🟡 Check Sales", int(counts.get("🟡 CHECK SALES", 0)))
    col4.metric("🟢 Normal", int(counts.get("🟢 NORMAL", 0)))

    if df["PO_Value_Rp"].sum() > 0:
        st.metric("Estimasi Nilai PO (harga pokok)", f"Rp {df['PO_Value_Rp'].sum():,.0f}")

    st.bar_chart(counts)

    st.subheader("Detail per SKU")

    filter_cols = st.columns(3)
    flag_filter = filter_cols[0].multiselect(
        "Filter status", options=list(FLAG_COLORS.keys()), default=list(FLAG_COLORS.keys())
    )
    brand_options = ["Semua"] + sorted(df[COLS["brand"]].dropna().unique().tolist()) \
        if COLS["brand"] in df.columns else ["Semua"]
    brand_filter = filter_cols[1].selectbox("Filter brand/jenis", brand_options)
    only_need_po = filter_cols[2].checkbox("Hanya tampilkan yang perlu PO", value=False)

    view = df[df["Flag"].isin(flag_filter)]
    if brand_filter != "Semua" and COLS["brand"] in df.columns:
        view = view[view[COLS["brand"]] == brand_filter]
    if only_need_po:
        view = view[view["PO_Qty"] > 0]

    display_cols = [c for c in [
        COLS["product_id"], COLS["nama_barang"], COLS.get("brand"),
        COLS["qty_terjual_30h"], COLS["stok"], "PO_Qty", "PO_Value_Rp", "Flag", "Alasan",
    ] if c in view.columns]

    def highlight_flag(row):
        color = FLAG_COLORS.get(row["Flag"])
        return [f"background-color: #{color}" if color else "" for _ in row]

    st.dataframe(
        view[display_cols].style.apply(highlight_flag, axis=1),
        use_container_width=True,
        height=450,
    )

    urgent_items = df[df["Flag"] == "🔴 URGENT"]
    if len(urgent_items) > 0:
        with st.expander(f"⚠️ {len(urgent_items)} item butuh perhatian segera"):
            st.dataframe(urgent_items[display_cols], use_container_width=True)

    # --------------------------------------------------------
    # EXCEL EXPORT (PO Value tidak diikutkan sesuai request)
    # --------------------------------------------------------
    def autosize_and_freeze(ws, dframe):
        ws.freeze_panes = "A2"
        for i, col in enumerate(dframe.columns, start=1):
            width = max(12, min(40, int(dframe[col].astype(str).str.len().max() or 10) + 2))
            ws.column_dimensions[get_column_letter(i)].width = width

    def write_sheet(wb, sheet_name, dframe, header_font, header_fill):
        safe_name = str(sheet_name)[:31] or "Sheet"
        ws = wb.create_sheet(safe_name)
        ws.append(list(dframe.columns))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        flag_idx = list(dframe.columns).index("Flag") + 1 if "Flag" in dframe.columns else None
        for _, row in dframe.iterrows():
            ws.append(list(row))
            if flag_idx:
                color = FLAG_COLORS.get(row["Flag"])
                if color:
                    ws.cell(row=ws.max_row, column=flag_idx).fill = PatternFill("solid", fgColor=color)
        autosize_and_freeze(ws, dframe)

    def build_excel_bytes(dframe) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)
        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="186156")  # brand teal

        ws = wb.create_sheet("Summary", 0)
        ws["A1"] = "PO Recommendation Summary - K. Beauty"
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color="186156")
        ws["A3"], ws["B3"] = "Total SKU", len(dframe)
        ws["A4"], ws["B4"] = "🔴 Urgent", int((dframe["Flag"] == "🔴 URGENT").sum())
        ws["A5"], ws["B5"] = "🟡 Check Sales", int((dframe["Flag"] == "🟡 CHECK SALES").sum())
        ws["A6"], ws["B6"] = "🟢 Normal", int((dframe["Flag"] == "🟢 NORMAL").sum())
        for col, width in zip("AB", (32, 20)):
            ws.column_dimensions[col].width = width

        # PO_Value_Rp sengaja tidak diikutkan di kolom export.
        all_cols = [c for c in [
            COLS["product_id"], COLS["nama_barang"], COLS.get("brand"),
            COLS["qty_terjual_30h"], COLS["stok"], "PO_Qty", "Flag", "Alasan",
        ] if c in dframe.columns]
        write_sheet(wb, "Semua SKU", dframe[all_cols], header_font, header_fill)

        need_po = dframe[dframe["PO_Qty"] > 0]
        if COLS["brand"] in dframe.columns and len(need_po) > 0:
            for brand, group in need_po.groupby(COLS["brand"]):
                write_sheet(wb, str(brand), group[all_cols], header_font, header_fill)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    st.divider()
    excel_bytes = build_excel_bytes(df)
    st.download_button(
        "⬇️ Download Excel (Summary + Semua SKU + per Brand)",
        data=excel_bytes,
        file_name="PO_Recommendation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ============================================================
# TAB 2: ANALITIK PER BRAND (minggu ini)
# ============================================================
with tab_brand:
    st.subheader("📦 Sales & Stock Analytics per Brand - Minggu Ini")

    if COLS["brand"] not in df.columns or df[COLS["brand"]].dropna().empty:
        st.caption("Kolom Brand/Jenis tidak ditemukan di data yang diupload.")
    else:
        brand_agg = df.groupby(COLS["brand"]).agg(
            Total_SKU=(COLS["product_id"], "count"),
            Qty_Terjual=(COLS["qty_terjual_30h"], "sum"),
            Total_Stok=(COLS["stok"], "sum"),
            PO_Qty=("PO_Qty", "sum"),
            Urgent=("Flag", lambda s: (s == "🔴 URGENT").sum()),
            Check_Sales=("Flag", lambda s: (s == "🟡 CHECK SALES").sum()),
        ).reset_index().rename(columns={COLS["brand"]: "Brand"})
        brand_agg = brand_agg.sort_values("Qty_Terjual", ascending=False)

        top_n = st.slider("Tampilkan berapa brand teratas", 5, min(40, len(brand_agg)),
                           value=min(15, len(brand_agg)))
        brand_view = brand_agg.head(top_n)

        chart_cols = st.columns(2)
        with chart_cols[0]:
            st.markdown("**Qty Terjual 30 Hari per Brand**")
            st.bar_chart(brand_view.set_index("Brand")["Qty_Terjual"], color=TEAL)
        with chart_cols[1]:
            st.markdown("**Total Stok per Brand**")
            st.bar_chart(brand_view.set_index("Brand")["Total_Stok"], color=LIME)

        st.markdown("**PO yang direkomendasikan per Brand**")
        st.bar_chart(brand_view.set_index("Brand")["PO_Qty"], color=TEAL)

        st.markdown("**Tabel lengkap per Brand**")
        st.dataframe(
            brand_agg.rename(columns={
                "Qty_Terjual": "Qty Terjual 30 Hari", "Total_Stok": "Total Stok",
                "PO_Qty": "PO Qty Direkomendasikan", "Check_Sales": "Check Sales",
            }),
            use_container_width=True, height=400,
        )
